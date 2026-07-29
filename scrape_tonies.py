"""
Scrape all Tonies retailers from the ChannelSight location API across the US.
Outputs a deduplicated, formatted Excel file (Tonies_Retailers_USA.xlsx).

Usage: python scrape_tonies.py
"""
import requests
import json
import time
from datetime import datetime
from collections import Counter
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR / "output"

API_BASE = "https://api.channelsight.com/api/2.42/location"
API_KEY = "ae24fee6-c62a-493a-b862-39186306cbf8"
ASSET_ID = "6391"
PID = "VSKU-tonies-512350"

CHAIN_INFO = {
    "Target": {
        "website": "https://www.target.com",
        "contact": "https://contactus.target.com",
    },
    "Walmart": {
        "website": "https://www.walmart.com",
        "contact": "https://corporate.walmart.com/contact",
    },
    "Kohls": {
        "website": "https://www.kohls.com",
        "contact": "https://cs.kohls.com/app/ask",
    },
    "Sam's Club": {
        "website": "https://www.samsclub.com",
        "contact": "https://help.samsclub.com/app/ask",
    },
    "Barnes & Noble": {
        "website": "https://www.barnesandnoble.com",
        "contact": "https://www.barnesandnoble.com/h/help",
    },
    "Meijer": {
        "website": "https://www.meijer.com",
        "contact": "https://www.meijer.com/content/corp/contact-us.html",
    },
    "Nordstrom": {
        "website": "https://www.nordstrom.com",
        "contact": "https://www.nordstrom.com/customer-service",
    },
    "Fred Meyer": {
        "website": "https://www.fredmeyer.com",
        "contact": "https://www.fredmeyer.com/hc/help",
    },
    "Toys 'R' Us @ Macy's": {
        "website": "https://www.macys.com/shop/toys",
        "contact": "https://customerservice.macys.com",
    },
    "Kroger": {
        "website": "https://www.kroger.com",
        "contact": "https://www.kroger.com/hc/help",
    },
    "Scheels": {
        "website": "https://www.scheels.com",
        "contact": "https://www.scheels.com/contact-us.html",
    },
}

METROS = [
    (40.7, -74.0), (34.0, -118.2), (41.9, -87.6), (29.8, -95.4),
    (33.4, -112.1), (40.0, -75.2), (29.4, -98.5), (32.7, -117.2),
    (32.8, -96.8), (37.3, -121.9), (30.3, -97.7), (30.3, -81.7),
    (35.2, -80.8), (37.8, -122.4), (39.7, -105.0), (47.6, -122.3),
    (38.9, -77.0), (36.2, -86.8), (42.4, -71.1), (25.8, -80.2),
    (33.7, -84.4), (45.0, -93.3), (36.2, -115.1), (28.0, -82.5),
    (38.3, -85.8), (35.1, -90.0), (43.0, -87.9), (39.1, -94.6),
    (35.5, -97.5), (35.8, -78.6), (41.3, -96.0), (21.3, -157.9),
    (61.2, -149.9), (26.1, -80.1), (27.3, -80.4), (28.5, -81.4),
    (26.6, -81.9), (42.3, -83.0), (39.3, -84.5), (40.4, -80.0),
    (43.1, -73.8), (40.7, -73.2), (41.1, -73.5), (43.6, -79.4),
    (39.0, -76.6), (37.5, -77.4), (36.8, -76.0), (35.8, -83.9),
    (34.7, -82.4), (32.8, -79.9), (33.5, -86.8), (30.4, -87.2),
    (30.0, -90.1), (44.5, -88.0), (43.0, -89.4), (41.7, -72.7),
    (41.8, -71.4), (43.2, -71.5), (44.5, -73.2), (42.9, -78.8),
    (43.0, -76.1), (33.8, -118.0), (34.2, -119.2), (33.1, -117.0),
    (38.6, -121.5), (36.7, -119.8), (34.4, -119.7), (46.9, -122.9),
    (45.5, -122.7), (44.1, -123.1), (40.8, -74.2), (40.2, -74.8),
    (39.4, -74.5), (41.5, -81.7), (40.0, -82.9), (41.1, -81.5),
    (39.8, -86.1), (41.7, -86.3), (40.4, -86.9), (42.7, -84.5),
    (42.3, -85.6), (43.0, -85.7), (46.8, -92.1), (44.0, -92.5),
    (41.6, -93.6), (42.5, -90.7), (38.6, -90.2), (38.8, -89.0),
    (40.1, -88.2), (39.8, -89.6), (41.5, -88.1), (42.0, -88.0),
]


def build_grid():
    points = []

    for lat in range(25, 50, 2):
        for lon in range(-125, -66, 2):
            points.append((lat, lon))

    for lat in range(55, 72, 3):
        for lon in range(-170, -130, 4):
            points.append((lat, lon))

    points.extend([(21.3, -157.8), (20.8, -156.3), (19.7, -155.5)])

    for clat, clon in METROS:
        for dlat in [-0.5, -0.25, 0, 0.25, 0.5]:
            for dlon in [-0.5, -0.25, 0, 0.25, 0.5]:
                points.append((clat + dlat, clon + dlon))

    seen = set()
    unique = []
    for lat, lon in points:
        key = (round(lat, 2), round(lon, 2))
        if key not in seen:
            seen.add(key)
            unique.append((lat, lon))

    return unique


def scrape_locations(grid):
    all_locations = {}
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": "https://us.tonies.com/",
    })

    for i, (lat, lon) in enumerate(grid):
        params = {
            "key": API_KEY,
            "assetId": ASSET_ID,
            "pid": PID,
            "lat": lat,
            "lon": lon,
        }
        try:
            resp = session.get(API_BASE, params=params, timeout=15)
            if resp.status_code == 200:
                locations = resp.json().get("Locations", [])
                new_count = sum(1 for loc in locations
                                if loc.get("LocationId") not in all_locations)
                for loc in locations:
                    loc_id = loc.get("LocationId")
                    if loc_id and loc_id not in all_locations:
                        all_locations[loc_id] = loc
                if (i + 1) % 100 == 0 or new_count > 0:
                    print(f"[{i+1}/{len(grid)}] ({lat:.1f},{lon:.1f}): "
                          f"{len(locations)} results, {new_count} new. "
                          f"Total: {len(all_locations)}")
            elif resp.status_code == 429:
                print(f"[{i+1}] Rate limited, waiting 5s...")
                time.sleep(5)
        except Exception as e:
            print(f"[{i+1}] Error: {e}")

        if i % 10 == 0:
            time.sleep(0.25)

    return all_locations


def build_excel(all_locations, output_path="Tonies_Retailers_USA.xlsx"):
    stores = []
    for loc in all_locations.values():
        name = (loc.get("RetailerLocationName")
                or loc.get("FriendlyName") or "")
        chain = CHAIN_INFO.get(name, {})

        stores.append({
            "Retailer Name": name,
            "Website URL": loc.get("StoreWebsite") or chain.get("website", ""),
            "Email Address": "",
            "Vendor/Wholesale Contact Form URL": chain.get("contact", ""),
            "City": loc.get("RetailerLocationCity", ""),
            "State": loc.get("RetailerLocationState", ""),
        })

    stores.sort(key=lambda x: (x["State"], x["City"], x["Retailer Name"]))

    seen = set()
    deduped = []
    for s in stores:
        key = (s["Retailer Name"], s["City"], s["State"])
        if key not in seen:
            seen.add(key)
            deduped.append(s)
    stores = deduped

    wb = Workbook()
    ws = wb.active
    ws.title = "Tonies Retailers USA"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C41E3A", end_color="C41E3A",
                              fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "Retailer Name", "Website URL", "Email Address",
        "Vendor/Wholesale Contact Form URL", "City", "State",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    link_font = Font(name="Calibri", color="0563C1", underline="single", size=11)
    url_columns = {"Website URL", "Vendor/Wholesale Contact Form URL"}

    for row_idx, store in enumerate(stores, 2):
        for col_idx, header in enumerate(headers, 1):
            value = store.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if header in url_columns and value.startswith("http"):
                cell.hyperlink = value
                cell.font = link_font

    col_widths = {"A": 35, "B": 40, "C": 30, "D": 50, "E": 22, "F": 8}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(stores) + 1}"

    wb.save(output_path)
    return stores


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    print("Building query grid...")
    grid = build_grid()
    print(f"Total grid points: {len(grid)}")

    print("\nScraping locations...")
    all_locations = scrape_locations(grid)
    print(f"\nTotal unique locations: {len(all_locations)}")

    json_path = OUTPUT_DIR / f"tonies_raw_{timestamp}.json"
    with open(json_path, "w") as f:
        json.dump(list(all_locations.values()), f, indent=2)
    print(f"Saved raw data to {json_path}")

    print("\nBuilding Excel...")
    excel_path = OUTPUT_DIR / f"Tonies_Retailers_USA_{timestamp}.xlsx"
    stores = build_excel(all_locations, output_path=str(excel_path))
    print(f"Saved {len(stores)} deduplicated retailers to {excel_path}")

    names = Counter(s["Retailer Name"] for s in stores)
    states = Counter(s["State"] for s in stores)
    print(f"\nUnique retailer brands: {len(names)}")
    print(f"States covered: {len(states)}")
    print("Top 10 retailers:")
    for name, count in names.most_common(10):
        print(f"  {name}: {count} locations")


if __name__ == "__main__":
    main()
