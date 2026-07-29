"""
Enrich Tonies retailer data with email addresses and websites.

For indie stores without websites: searches Google to find their website.
For all stores with websites: scrapes the site for email addresses.

Usage: python enrich_emails.py <path_to_raw_json>
  e.g.: python enrich_emails.py output/tonies_raw_2026-07-19_21-30-45.json

Outputs an enriched JSON and rebuilt Excel in the output/ folder.
"""
import re
import sys
import json
import time
import requests
from datetime import datetime
from pathlib import Path
from collections import Counter
from urllib.parse import urlparse, unquote
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR / "output"

EMAIL_REGEX = re.compile(
    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
)

JUNK_EMAILS = {
    "example.com", "example.org", "sentry.io", "wixpress.com",
    "sentry-next.wixpress.com", "googleapis.com", "google.com",
    "facebook.com", "twitter.com", "instagram.com", "youtube.com",
    "schema.org", "w3.org", "apple.com", "microsoft.com",
    "shopify.com", "wordpress.com", "squarespace.com", "wix.com",
    "godaddy.com", "cloudflare.com", "email.com", "company.site",
    "birdeye.com", "sentry.wixpress.com",
}

CHAIN_INFO = {
    "Target": {
        "website": "https://www.target.com",
        "contact": "https://contactus.target.com",
    },
    "Walmart": {
        "website": "https://www.walmart.com",
        "contact": "https://corporate.walmart.com/contact",
    },
    "Kohls": {
        "website": "https://www.kohls.com",
        "contact": "https://cs.kohls.com/app/ask",
    },
    "Sam's Club": {
        "website": "https://www.samsclub.com",
        "contact": "https://help.samsclub.com/app/ask",
    },
    "Barnes & Noble": {
        "website": "https://www.barnesandnoble.com",
        "contact": "https://www.barnesandnoble.com/h/help",
    },
    "Meijer": {
        "website": "https://www.meijer.com",
        "contact": "https://www.meijer.com/content/corp/contact-us.html",
    },
    "Nordstrom": {
        "website": "https://www.nordstrom.com",
        "contact": "https://www.nordstrom.com/customer-service",
    },
    "Fred Meyer": {
        "website": "https://www.fredmeyer.com",
        "contact": "https://www.fredmeyer.com/hc/help",
    },
    "Toys 'R' Us @ Macy's": {
        "website": "https://www.macys.com/shop/toys",
        "contact": "https://customerservice.macys.com",
    },
    "Kroger": {
        "website": "https://www.kroger.com",
        "contact": "https://www.kroger.com/hc/help",
    },
    "Scheels": {
        "website": "https://www.scheels.com",
        "contact": "https://www.scheels.com/contact-us.html",
    },
}

CONTACT_PAGE_PATHS = [
    "/contact", "/contact-us", "/contactus", "/contact.html",
    "/pages/contact", "/pages/contact-us",
    "/about/contact", "/about-us/contact",
    "/reach-us", "/get-in-touch",
]

session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
})


def is_junk_email(email):
    domain = email.split("@")[1].lower()
    if domain in JUNK_EMAILS:
        return True
    if "noreply" in email.lower() or "no-reply" in email.lower():
        return True
    if "unsubscribe" in email.lower() or "mailer" in email.lower():
        return True
    if email.endswith(".png") or email.endswith(".jpg"):
        return True
    if email.endswith(".gif") or email.endswith(".webp") or email.endswith(".css"):
        return True
    if "@" not in email or "." not in email.split("@")[1]:
        return True
    return False


def extract_emails_from_html(html):
    raw_emails = EMAIL_REGEX.findall(html)
    cleaned = set()
    for email in raw_emails:
        email = email.lower().strip(".")
        if not is_junk_email(email):
            cleaned.add(email)
    return list(cleaned)


def scrape_website_for_email(url):
    """Visit a website's homepage and contact page, extract email addresses."""
    if not url or not url.startswith("http"):
        return [], ""

    emails = set()
    contact_url = ""

    try:
        resp = session.get(url, timeout=10, allow_redirects=True)
        if resp.status_code == 200:
            found = extract_emails_from_html(resp.text)
            emails.update(found)

            # Look for contact page link in the HTML
            contact_match = re.search(
                r'href=["\']([^"\']*(?:contact|reach|get-in-touch)[^"\']*)["\']',
                resp.text, re.IGNORECASE
            )
            if contact_match:
                contact_path = contact_match.group(1)
                if contact_path.startswith("http"):
                    contact_url = contact_path
                elif contact_path.startswith("/"):
                    parsed = urlparse(url)
                    contact_url = f"{parsed.scheme}://{parsed.netloc}{contact_path}"
    except Exception:
        pass

    # Also try common contact page paths
    if not emails:
        parsed = urlparse(url)
        base = f"{parsed.scheme}://{parsed.netloc}"
        for path in CONTACT_PAGE_PATHS:
            try:
                resp = session.get(base + path, timeout=8, allow_redirects=True)
                if resp.status_code == 200:
                    found = extract_emails_from_html(resp.text)
                    if found:
                        emails.update(found)
                        if not contact_url:
                            contact_url = base + path
                        break
            except Exception:
                continue
            time.sleep(0.3)

    return list(emails), contact_url


def search_for_website(store_name, city, state):
    """Search DuckDuckGo for a store's website."""
    query = f"{store_name} {city} {state}"
    try:
        resp = session.get(
            "https://html.duckduckgo.com/html/",
            params={"q": query},
            headers={"Referer": "https://duckduckgo.com/"},
            timeout=10,
        )
        if resp.status_code == 200:
            skip = [
                "duckduckgo.", "google.", "facebook.", "instagram.",
                "twitter.", "yelp.", "yellowpages.", "mapquest.",
                "tripadvisor.", "bbb.", "linkedin.", "pinterest.",
                "tiktok.", "amazon.", "walmart.", "target.", "youtube.",
            ]
            for link in re.findall(r'href="([^"]+)"', resp.text):
                if "uddg=" in link:
                    url = unquote(link.split("uddg=")[1].split("&")[0])
                    domain = urlparse(url).netloc.lower()
                    if not any(s in domain for s in skip):
                        return url
    except Exception:
        pass
    return ""


def build_excel(stores, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tonies Retailers USA"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C41E3A", end_color="C41E3A",
                              fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "Retailer Name", "Website URL", "Email Address",
        "Vendor/Wholesale Contact Form URL", "City", "State",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    link_font = Font(name="Calibri", color="0563C1", underline="single", size=11)
    url_columns = {"Website URL", "Vendor/Wholesale Contact Form URL"}

    for row_idx, store in enumerate(stores, 2):
        for col_idx, header in enumerate(headers, 1):
            value = store.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if header in url_columns and value.startswith("http"):
                cell.hyperlink = value
                cell.font = link_font

    col_widths = {"A": 35, "B": 40, "C": 30, "D": 50, "E": 22, "F": 8}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(stores) + 1}"
    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        # Find the most recent raw JSON in the output folder
        json_files = sorted(OUTPUT_DIR.glob("tonies_raw_*.json"))
        if not json_files:
            print("No raw JSON found. Run scrape_tonies.py first.")
            return
        input_path = json_files[-1]
        print(f"Using most recent: {input_path}")
    else:
        input_path = Path(sys.argv[1])

    with open(input_path) as f:
        data = json.load(f)

    # Deduplicate into unique retailers (by name + city + state)
    seen = {}
    for loc in data:
        name = loc.get("RetailerLocationName") or loc.get("FriendlyName") or ""
        city = loc.get("RetailerLocationCity", "")
        state = loc.get("RetailerLocationState", "")
        key = (name, city, state)
        if key not in seen:
            seen[key] = loc

    print(f"Unique retailers: {len(seen)}")

    # Identify indie stores (not major chains)
    chain_names = set(CHAIN_INFO.keys())
    indie_keys = [k for k in seen if k[0] not in chain_names]
    print(f"Indie stores to enrich: {len(indie_keys)}")

    # Step 1: Find websites for indie stores via Google
    print("\n--- Step 1: Finding indie store websites ---")
    indie_websites = {}
    for i, key in enumerate(indie_keys):
        name, city, state = key
        website = search_for_website(name, city, state)
        if website:
            indie_websites[key] = website
            print(f"  [{i+1}/{len(indie_keys)}] {name} ({city}, {state}): {website}")
        else:
            print(f"  [{i+1}/{len(indie_keys)}] {name} ({city}, {state}): not found")
        time.sleep(1.5)  # Be polite to Google

    print(f"\nFound websites for {len(indie_websites)}/{len(indie_keys)} indie stores")

    # Step 2: Scrape emails from all websites (chains + indies)
    print("\n--- Step 2: Scraping emails from websites ---")
    email_results = {}
    contact_results = {}

    # Scrape indie stores
    for i, key in enumerate(indie_keys):
        name, city, state = key
        url = indie_websites.get(key, "")
        if url:
            emails, contact_url = scrape_website_for_email(url)
            if emails:
                email_results[key] = emails[0]
                print(f"  [{i+1}/{len(indie_keys)}] {name}: {emails[0]}")
            if contact_url:
                contact_results[key] = contact_url
            time.sleep(0.5)

    print(f"\nFound emails for {len(email_results)} indie stores")
    print(f"Found contact pages for {len(contact_results)} indie stores")

    # Step 3: Build final store list
    print("\n--- Step 3: Building final Excel ---")
    stores = []
    for key, loc in seen.items():
        name, city, state = key
        chain = CHAIN_INFO.get(name, {})

        website = (loc.get("StoreWebsite")
                   or indie_websites.get(key, "")
                   or chain.get("website", ""))
        email = email_results.get(key, "")
        contact = contact_results.get(key, "") or chain.get("contact", "")

        stores.append({
            "Retailer Name": name,
            "Website URL": website,
            "Email Address": email,
            "Vendor/Wholesale Contact Form URL": contact,
            "City": city,
            "State": state,
        })

    stores.sort(key=lambda x: (x["State"], x["City"], x["Retailer Name"]))

    OUTPUT_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    # Save enriched JSON
    json_path = OUTPUT_DIR / f"tonies_enriched_{timestamp}.json"
    with open(json_path, "w") as f:
        json.dump(stores, f, indent=2)
    print(f"Saved enriched JSON to {json_path}")

    # Save Excel
    excel_path = OUTPUT_DIR / f"Tonies_Retailers_USA_{timestamp}.xlsx"
    build_excel(stores, str(excel_path))
    print(f"Saved {len(stores)} retailers to {excel_path}")

    # Summary
    has_website = sum(1 for s in stores if s["Website URL"])
    has_email = sum(1 for s in stores if s["Email Address"])
    has_contact = sum(1 for s in stores if s["Vendor/Wholesale Contact Form URL"])
    print(f"\nCoverage:")
    print(f"  Website URL: {has_website}/{len(stores)}")
    print(f"  Email Address: {has_email}/{len(stores)}")
    print(f"  Contact Form: {has_contact}/{len(stores)}")


if __name__ == "__main__":
    main()
